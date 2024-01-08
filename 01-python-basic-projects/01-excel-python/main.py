from openpyxl import load_workbook
#from openpyxl import Workbook
work_book = load_workbook('test.xlsx')
ws = work_book.active
file_name = ws['B2'].value
print(file_name)
work_book.save("{}.xlsx".format(file_name))

#print(book.sheetnames)

#wb = Workbook()
#ws = wb.active

#file_name = wb['Sheet1']['B2']
#print(file_name)
#wb.save("test2.xlsx")
#ws['A1'] = 'hello world'
#wb.save("{}x.xlsx".format(file_name))

#ws['A1'] = 42
#
#ws.append([1,2,3])

#import datetime
#ws['A2'] = datetime.datetime.now()
#wb.save("sample.xlsx")*/
